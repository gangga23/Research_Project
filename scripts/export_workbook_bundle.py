"""
Assemble CSV/text/XLSX deliverables from ``app_master`` + ``app_version_history`` frames.

Used after scraping (``run_pipeline.py``) or offline from cached CSVs (``build_workbook_only.py``).
"""

from __future__ import annotations

import importlib.util
import math
import sys
from pathlib import Path

import pandas as pd

_SPOTIFY_CONTAM_RE = r"(?:With the Spotify music and podcast app|WHY SPOTIFY FOR MUSIC AND PODCASTS\?)"


def _fix_mojibake_release_notes(version_df: pd.DataFrame) -> pd.DataFrame:
    """
    Repair common mojibake punctuation in cached release_notes without re-scraping
    (e.g. "Weâ€™re" -> "We’re").
    """
    if version_df.empty or "release_notes" not in version_df.columns:
        return version_df
    rn = version_df["release_notes"].fillna("").astype(str)
    if not rn.str.contains(r"[âÃ]", regex=True).any():
        return version_df
    df = version_df.copy()
    df["release_notes"] = (
        rn.str.replace("â€™", "’", regex=False)
        .str.replace("â€˜", "‘", regex=False)
        .str.replace("â€œ", "“", regex=False)
        .str.replace("â€�", "”", regex=False)
        .str.replace("â€“", "–", regex=False)
        .str.replace("â€”", "—", regex=False)
        .str.replace("â€¦", "…", regex=False)
        .str.replace("Â ", " ", regex=False)
        .str.replace("Â", "", regex=False)
    )
    return df
def _relabel_update_category(version_df: pd.DataFrame) -> pd.DataFrame:
    """
    Recompute update_category from release_notes using current CATEGORY_RULES.

    Intent: reduce "Other" without re-scraping, and correct obvious over-labeling
    caused by earlier, too-broad regexes (Android wayback/store text is especially noisy).
    """
    if version_df.empty:
        return version_df
    if "release_notes" not in version_df.columns:
        return version_df

    import run_pipeline as rp

    df = version_df.copy()
    df["update_category"] = df["release_notes"].fillna("").astype(str).apply(rp.pick_update_category)
    return df


def _sanitize_android_release_notes_contamination(version_df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """
    Defensive cleanup for cross-app scraper contamination observed in some exports where
    multiple Android rows share Spotify's Play description as release_notes.

    Sets contaminated Android release_notes to "Not available" and update_category to "Other".
    """
    if version_df.empty:
        return version_df, 0
    if not {"platform", "release_notes", "update_category"}.issubset(set(version_df.columns)):
        return version_df, 0

    df = version_df.copy()
    df["_dq_pipeline_note"] = ""
    rn = df["release_notes"].fillna("").astype(str)
    mask = (df["platform"].astype(str) == "Android") & rn.str.contains(_SPOTIFY_CONTAM_RE, case=False, regex=True)
    n = int(mask.sum())
    if n:
        df.loc[mask, "release_notes"] = "Not available"
        df.loc[mask, "update_category"] = "Other"
        df.loc[mask, "_dq_pipeline_note"] = (
            "release_notes cleared after scrape matched known cross-app contamination pattern "
            "(Spotify marketing copy on non-Spotify listings)."
        )
    return df, n


def _load_submission_summary(script_dir: Path):
    sd = str(script_dir.resolve())
    if sd not in sys.path:
        sys.path.insert(0, sd)
    path = script_dir / "submission_summary.py"
    if not path.is_file():
        raise ImportError(f"Missing required file: {path}")
    spec = importlib.util.spec_from_file_location("submission_summary", path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Cannot load module spec for {path}")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def infer_both_platform_app_count(master_df: pd.DataFrame) -> int:
    """Count app_name keys present on both iOS and Android in ``app_master``."""
    if master_df.empty or "app_name" not in master_df.columns or "platform" not in master_df.columns:
        return 0

    def _plat_set(g: pd.DataFrame) -> set[str]:
        return set(g["platform"].astype(str).unique())

    by_app = master_df.groupby("app_name", sort=False)
    n = 0
    for _, g in by_app:
        s = _plat_set(g)
        if "iOS" in s and "Android" in s:
            n += 1
    return n


def validate_frames(master_df: pd.DataFrame, version_df: pd.DataFrame) -> None:
    """Raise RuntimeError if enums/columns are inconsistent."""
    import run_pipeline as rp  # Local import avoids circular load at package init.

    req_m = {
        "app_id",
        "app_name",
        "platform",
        "developer",
        "category",
        "initial_release_date",
        "source_url",
        "current_version",
        "current_version_release_date",
        "notes",
    }
    req_v = {
        "app_id",
        "app_name",
        "platform",
        "version_number",
        "release_date",
        "release_notes",
        "source_type",
        "confidence_level",
        "update_category",
    }
    miss_m = req_m - set(master_df.columns)
    miss_v = req_v - set(version_df.columns)
    if miss_m:
        raise RuntimeError(f"app_master.csv missing columns: {miss_m}")
    if miss_v:
        raise RuntimeError(f"app_version_history.csv missing columns: {miss_v}")

    bad = set(version_df["update_category"].unique()) - set(rp.UPDATE_CATEGORIES)
    if bad:
        raise RuntimeError(f"Invalid update_category values: {bad}")
    bad_s = set(version_df["source_type"].unique()) - rp.ALLOWED_SOURCE_TYPES
    if bad_s:
        raise RuntimeError(f"Invalid source_type values: {bad_s}")
    bad_c = set(version_df["confidence_level"].unique()) - rp.ALLOWED_CONFIDENCE
    if bad_c:
        raise RuntimeError(f"Invalid confidence_level values: {bad_c}")


def _apply_normalized_workbook_openpyxl_formatting(xlsx_path: Path) -> None:
    """
    Post-write styling for ``normalized_dataset.xlsx``: fonts, fills, freeze panes,
    widths, wrap, alternating rows, autofilter, tab colors, gridline visibility.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.styles.colors import Color
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    FONT = Font(name="Calibri", size=11)
    FONT_BOLD = Font(name="Calibri", size=11, bold=True)
    FONT_LINK = Font(name="Calibri", size=11, color="0563C1", underline="single")
    FILL_HDR = PatternFill(fill_type="solid", start_color="DAEEF3", end_color="DAEEF3")
    FILL_WHITE = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
    FILL_ALT = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
    AL_WRAP_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
    AL_NOWRAP_TOP = Alignment(horizontal="left", vertical="top", wrap_text=False)

    def _rgb_tab(rgb_hex: str) -> Color:
        s = rgb_hex.strip().lstrip("#").upper()
        if len(s) == 6:
            s = "FF" + s
        return Color(rgb=s)

    TAB_COLORS = {
        "submission_observations": _rgb_tab("4472C4"),
        "app_master": _rgb_tab("70AD47"),
        "submission_summary": _rgb_tab("ED7D31"),
        "viz_fast_scan": _rgb_tab("7030A0"),
    }

    def _max_line_len(val: object) -> int:
        if val is None:
            return 0
        return max((len(line) for line in str(val).splitlines()), default=0)

    def _col_display_width(ws, col_idx: int, *, cap: float | None) -> float:
        m = 10
        for r in range(1, ws.max_row + 1):
            m = max(m, _max_line_len(ws.cell(r, col_idx).value))
        w = min(m + 2.5, 120.0 if cap is None else float(cap))
        return max(9.0, w)

    def _submission_summary_row_height(val: object, *, col_width: float = 85.0) -> float:
        """Approximate Excel row height (points) for wrapped text at ``col_width`` character columns."""
        if val is None:
            return 15.0
        s = str(val)
        chars_per_line = max(int(col_width * 0.85), 24)
        lines = 1
        for para in s.splitlines():
            lines += max(1, math.ceil(len(para) / chars_per_line))
        return float(min(400.0, max(15.0, 12.0 + lines * 13.5)))

    wb = load_workbook(xlsx_path)

    # --- submission_observations ---
    if "submission_observations" in wb.sheetnames:
        ws = wb["submission_observations"]
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = True
        hdr_map: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            if v is not None:
                hdr_map[str(v)] = c
        rn_ci = hdr_map.get("release_notes")
        notes_ci = hdr_map.get("notes")
        summ_ci = hdr_map.get("update_summary")
        url_cols_ci = [hdr_map.get("history_source_url")]

        for c in range(1, ws.max_column + 1):
            cap = 60.0 if c in {rn_ci, notes_ci, summ_ci} else None
            ws.column_dimensions[get_column_letter(c)].width = _col_display_width(ws, c, cap=cap)

        def _dim_width(col_idx: int | None, *, default: float = 50.0) -> float:
            if col_idx is None:
                return default
            dim = ws.column_dimensions[get_column_letter(col_idx)].width
            return float(dim) if dim is not None else default

        w_rn = _dim_width(rn_ci)
        w_notes = _dim_width(notes_ci)
        w_summ = _dim_width(summ_ci, default=48.0)

        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                cell.font = FONT_BOLD if r == 1 else FONT
                if r == 1:
                    cell.fill = FILL_HDR
                    cell.alignment = AL_NOWRAP_TOP
                else:
                    cell.fill = FILL_WHITE if r % 2 == 0 else FILL_ALT
                    if c == rn_ci or c == notes_ci or c == summ_ci:
                        cell.alignment = AL_WRAP_TOP
                    else:
                        cell.alignment = AL_NOWRAP_TOP

        if ws.max_row >= 2:
            for r in range(2, ws.max_row + 1):
                h = 45.0
                if rn_ci is not None:
                    h = max(
                        h,
                        _submission_summary_row_height(ws.cell(r, rn_ci).value, col_width=w_rn),
                    )
                if notes_ci is not None:
                    h = max(
                        h,
                        _submission_summary_row_height(ws.cell(r, notes_ci).value, col_width=w_notes),
                    )
                if summ_ci is not None:
                    h = max(
                        h,
                        _submission_summary_row_height(ws.cell(r, summ_ci).value, col_width=w_summ),
                    )
                ws.row_dimensions[r].height = min(float(h), 240.0)

        for url_ci in url_cols_ci:
            if url_ci is None:
                continue
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(r, url_ci)
                raw = cell.value
                if raw is None:
                    continue
                url = str(raw).strip()
                if url.startswith(("http://", "https://")):
                    cell.hyperlink = url
                    cell.value = url
                    cell.font = FONT_LINK

        if ws.max_row >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # --- app_master ---
    if "app_master" in wb.sheetnames:
        ws = wb["app_master"]
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = True
        hdr_am: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            hv = ws.cell(1, c).value
            if hv is not None:
                hdr_am[str(hv)] = c
        master_url_ci = hdr_am.get("source_url")
        notes_am_ci = hdr_am.get("notes")

        for c in range(1, ws.max_column + 1):
            cap_w = 56.0 if c == notes_am_ci else 40.0
            w = min(_col_display_width(ws, c, cap=None), cap_w)
            ws.column_dimensions[get_column_letter(c)].width = w

        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                cell.font = FONT_BOLD if r == 1 else FONT
                if r == 1:
                    cell.fill = FILL_HDR
                    cell.alignment = AL_NOWRAP_TOP
                elif c == notes_am_ci:
                    cell.alignment = AL_WRAP_TOP
                else:
                    cell.alignment = AL_NOWRAP_TOP

        if master_url_ci is not None:
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(r, master_url_ci)
                raw = cell.value
                if raw is None:
                    continue
                url = str(raw).strip()
                if url.startswith(("http://", "https://")):
                    cell.hyperlink = url
                    cell.value = url
                    cell.font = FONT_LINK

        if ws.max_row >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # --- submission_summary ---
    if "submission_summary" in wb.sheetnames:
        ws = wb["submission_summary"]
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 28.0
        ws.column_dimensions["B"].width = 85.0
        for r in range(1, ws.max_row + 1):
            ha = ws.cell(r, 1)
            hb = ws.cell(r, 2)
            ha.font = FONT_BOLD
            ha.fill = FILL_HDR
            ha.alignment = AL_NOWRAP_TOP
            hb.font = FONT
            hb.alignment = AL_WRAP_TOP
            ws.row_dimensions[r].height = max(
                _submission_summary_row_height(hb.value, col_width=85.0),
                _submission_summary_row_height(ha.value, col_width=28.0),
            )

    # --- viz_fast_scan (visualizations) ---
    if "viz_fast_scan" in wb.sheetnames:
        ws = wb["viz_fast_scan"]
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                cell.font = FONT
                if isinstance(cell.value, str) and "\n" in cell.value:
                    cell.alignment = AL_WRAP_TOP
                else:
                    cell.alignment = AL_NOWRAP_TOP

    for sheet_name, tc in TAB_COLORS.items():
        if sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_properties.tabColor = tc

    wb.save(xlsx_path)


def export_workbook_bundle(
    master_df: pd.DataFrame,
    version_df: pd.DataFrame,
    *,
    n_config_apps: int,
    both_platforms: int,
    feed_validations: list,
    output_dir: Path,
    repo_root: Path,
    script_dir: Path,
    rewrite_master_version_csv: bool = True,
    rewrite_feed_validation_report: bool = True,
) -> str:
    """
    Write schema txt, CSVs (optional rewrite of master/version), reports, Excel, styling.

    Returns the validation_report text for printing.
    """
    import run_pipeline as rp  # For reporting helpers + schema_text.

    version_df = _fix_mojibake_release_notes(version_df)
    version_df, contam_n = _sanitize_android_release_notes_contamination(version_df)
    version_df = _relabel_update_category(version_df)
    version_df = version_df.copy()
    if "history_source_url" not in version_df.columns:
        version_df["history_source_url"] = ""
    else:
        version_df["history_source_url"] = (
            version_df["history_source_url"].fillna("").astype(str).map(lambda x: x.strip())
        )
    try:
        from apkmirror_history_merge import merge_apkmirror_history_urls

        version_df = merge_apkmirror_history_urls(version_df, repo_root / "data" / "cache")
    except ImportError:
        pass
    validate_frames(master_df, version_df)
    output_dir.mkdir(parents=True, exist_ok=True)

    ss = _load_submission_summary(script_dir)
    repo_url = ss.load_repository_url(repo_root)

    submission_obs_df = ss.build_submission_observations(version_df, master_df)

    version_df_public = version_df.drop(columns=["_dq_pipeline_note"], errors="ignore")

    (output_dir / "schema_tables.txt").write_text(rp.schema_text(), encoding="utf-8")
    if rewrite_master_version_csv:
        master_df.to_csv(output_dir / "app_master.csv", index=False, encoding="utf-8")
        version_df_public.to_csv(output_dir / "app_version_history.csv", index=False, encoding="utf-8")
    elif contam_n:
        # Even in "rebuild from cache" mode, we must not leave contaminated cached CSVs in place.
        version_df_public.to_csv(output_dir / "app_version_history.csv", index=False, encoding="utf-8")
        print(
            f"[warn] Sanitized {contam_n} contaminated Android release_notes rows and rewrote app_version_history.csv.",
            file=sys.stderr,
        )
    submission_obs_df.to_csv(output_dir / "submission_observations.csv", index=False, encoding="utf-8")

    rep = rp.validation_report(
        n_config_apps, len(master_df), len(version_df_public), version_df_public, both_platforms
    )
    (output_dir / "validation_report.txt").write_text(rep + "\n", encoding="utf-8")
    (output_dir / "data_quality_report.txt").write_text(
        rp.data_quality_report(version_df_public, n_config_apps, len(master_df)) + "\n", encoding="utf-8"
    )
    feed_path = output_dir / "feed_validation_report.txt"
    if rewrite_feed_validation_report or not feed_path.is_file():
        feed_path.write_text(rp.format_feed_validation_report(feed_validations), encoding="utf-8")

    metrics = []
    for line in rep.split("\n"):
        if ":" in line:
            k, v = line.split(":", 1)
            metrics.append({"metric": k.strip(), "value": v.strip()})
        elif line.strip():
            metrics.append({"metric": line.strip(), "value": ""})
    val_df = pd.DataFrame(metrics)
    dq_txt = rp.data_quality_report(version_df_public, n_config_apps, len(master_df))
    dq_metrics = []
    for line in dq_txt.split("\n"):
        if ":" in line:
            k, v = line.split(":", 1)
            dq_metrics.append({"metric": k.strip(), "value": v.strip()})
        elif line.strip():
            dq_metrics.append({"metric": line.strip(), "value": ""})
    submission_df = ss.build_submission_summary_dataframe(
        version_df_public,
        n_config_apps=n_config_apps,
        repo_url=repo_url,
        validation_text=rep,
        data_quality_text=dq_txt,
    )

    xlsx = output_dir / "normalized_dataset.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as wr:
        submission_obs_df.to_excel(wr, sheet_name="submission_observations", index=False)
        master_df.to_excel(wr, sheet_name="app_master", index=False)
        submission_df.to_excel(wr, sheet_name="submission_summary", index=False)

    try:
        from visualization_summary import try_append_visualization_sheet

        try_append_visualization_sheet(xlsx, version_df_public)
    except Exception as e:
        print(f"[warn] viz_fast_scan hook failed: {e}", file=sys.stderr)

    _apply_normalized_workbook_openpyxl_formatting(xlsx)

    return rep
