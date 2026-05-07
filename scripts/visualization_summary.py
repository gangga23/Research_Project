"""
Fast-scan visuals aligned with ``timeseries_insights_core`` (same facts as
``submission_summary.build_timeseries_insights``):

This sheet is intentionally small and rubric-aligned:
- update frequency over time by app (cadence heatmaps — iOS / Android)
- iOS vs Android observation depth per app (dated span)
- category share shift (oldest vs newest quartile; slope chart)

Uses Matplotlib (Agg) + openpyxl images; no network.
"""

from __future__ import annotations

import functools
import io
import math
import re
import sys
from pathlib import Path
from typing import Any

import pandas as pd

from timeseries_insights_core import (
    build_automated_trend_synopsis_bullets,
    dated_subset,
    parse_release_dates,
)
from version_display import version_string_missing

_SHEET = "charts"

# Cadence heatmaps (iOS/Android): cap color normalization so 0–20 uses full colormap;
# counts above the cap share the top color (colorbar ``extend='max'``).
CADENCE_HEATMAP_VMAX_CAP = 20.0


@functools.lru_cache(maxsize=1)
def _cadence_heatmap_colormap():
    """White at zero, then yellow → blue (no green band). Matplotlib only when charts run."""
    from matplotlib.colors import LinearSegmentedColormap

    return LinearSegmentedColormap.from_list(
        "cadence_yb",
        [
            "#ffffff",
            "#fffef7",
            "#fff9c4",
            "#ffeb3b",
            "#c5e1f5",
            "#64b5f6",
            "#1976d2",
            "#0d47a1",
        ],
        N=256,
    )


def build_automated_trend_synopsis(version_df: pd.DataFrame) -> list[str]:
    """Bullets from shared core (coverage → cadence → quartiles → strategy read → epistemic)."""
    # Keep the viz sheet synopsis short/scannable.
    return build_automated_trend_synopsis_bullets(version_df)[:6]


def _save_png_bytes(buf: io.BytesIO, out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_bytes(buf.getvalue())
    return out_path


def _short_quarter_tick_label(col: str) -> str:
    """Human-readable quarter ticks from pandas Period strings like ``2024Q1``."""
    s = str(col).strip()
    m = re.match(r"^(\d{4})Q([1-4])$", s)
    if m:
        return f"Q{m.group(2)} '{m.group(1)[2:]}"
    return s


def _compact_month_tick_label(ym: str) -> str:
    """Turn ``YYYY-MM`` period strings into short ticks like ``May '23``."""
    s = str(ym).strip()
    try:
        ts = pd.Timestamp(s)
        return f"{ts.strftime('%b')} '{ts.strftime('%y')}"
    except (ValueError, TypeError):
        return s


def _heatmap_xtick_positions(n: int, *, target_labels: int = 13) -> list[int]:
    """Sparse tick indices so month axes stay readable when many columns are drawn."""
    if n <= 0:
        return []
    if n <= target_labels:
        return list(range(n))
    stride = max(1, math.ceil(n / target_labels))
    out = list(range(0, n, stride))
    if out[-1] != n - 1:
        out.append(n - 1)
    return sorted(set(out))


def _cadence_heatmap_prepare(
    version_df: pd.DataFrame,
    *,
    platform: str,
    app_order: list[str],
    max_bins: int | None,
    bin_period: str,
    since: pd.Timestamp | None = None,
    fill_period_bins_since: pd.Timestamp | None = None,
    fill_end: pd.Timestamp | None = None,
) -> tuple[Any, list[str], list[str]] | None:
    """
    Build cadence matrix for ``imshow``: rows = apps (shared order), cols = time bins.
    Returns ``(mat, x_tick_labels, row_labels)`` or ``None``.

    ``since`` drops dated rows before that timestamp (platform-specific windows).

    ``fill_period_bins_since`` reindexes to every calendar bin from that anchor through
    the latest dated row (zeros where nothing was captured). Supports ``bin_period``
    ``M`` or ``Q`` (and ``Y``).

    When ``fill_end`` is set with ``fill_period_bins_since``, the reindex uses that end
    (e.g. global max date) so iOS/Android monthly matrices share the same column range.
    """
    sub = dated_subset(version_df)
    if sub is None:
        return None
    sub = sub[sub["platform"].astype(str) == platform]
    if len(sub) < 1:
        return None

    sub = sub.copy()
    if since is not None:
        sub = sub[sub["_dt"] >= since]
    if len(sub) < 1:
        return None

    sub["_bin"] = sub["_dt"].dt.to_period(bin_period).astype(str)
    piv = sub.pivot_table(index="app_name", columns="_bin", values="app_id", aggfunc="count", fill_value=0)
    if piv.empty:
        return None

    try:
        ordered_cols = sorted(piv.columns, key=lambda c: pd.Period(str(c), freq=bin_period))
    except (ValueError, TypeError):
        ordered_cols = list(piv.columns)
    piv = piv.reindex(columns=ordered_cols)

    if fill_period_bins_since is not None:
        anchor = pd.Timestamp(fill_period_bins_since)
        end_dt = sub["_dt"].max()
        if pd.isna(end_dt):
            return None
        end_ts = pd.Timestamp(fill_end) if fill_end is not None else pd.Timestamp(end_dt)
        freq_map = {"M": "M", "Q": "Q", "Y": "Y"}
        freq = freq_map.get(bin_period)
        if freq is None:
            pass
        else:
            start_p = anchor.to_period(freq)
            end_p = end_ts.to_period(freq)
            if end_p >= start_p:
                full_cols = pd.period_range(start_p, end_p, freq=freq).astype(str)
                piv = piv.reindex(columns=list(full_cols), fill_value=0)

    cols = list(piv.columns)
    if max_bins is not None and len(cols) > max_bins:
        piv = piv[cols[-max_bins:]]

    piv = piv.reindex(app_order).fillna(0)

    apps = list(piv.index.astype(str))
    raw_cols = list(piv.columns.astype(str))
    if bin_period == "Q":
        xlabels = [_short_quarter_tick_label(c) for c in raw_cols]
    elif bin_period == "M":
        xlabels = [_compact_month_tick_label(c) for c in raw_cols]
    else:
        xlabels = [str(c) for c in raw_cols]
    mat = piv.values.astype(float)
    return mat, xlabels, apps


def _chart_update_frequency_heatmap_platform(
    version_df: pd.DataFrame,
    *,
    platform: str,
    app_order: list[str],
    max_bins: int | None,
    bin_period: str,
    title: str,
    subtitle: str,
    vmax: float | None = None,
    prepared: tuple[Any, list[str], list[str]] | None = None,
    since: pd.Timestamp | None = None,
    fill_period_bins_since: pd.Timestamp | None = None,
    fill_end: pd.Timestamp | None = None,
    cmap=None,
) -> io.BytesIO | None:
    """
    Cadence heatmap for a single platform (dated rows).
    Rows: app_name; columns: YYYY-MM or YYYYQn; values: observation counts.
    Use the same ``vmax`` and default colormap for iOS and Android so numeric and hue scaling match.
    ``vmax`` is normally capped (see ``CADENCE_HEATMAP_VMAX_CAP``) so low/mid counts show more hue spread.
    """
    import matplotlib.pyplot as plt
    import numpy as np

    block = prepared
    if block is None:
        block = _cadence_heatmap_prepare(
            version_df,
            platform=platform,
            app_order=app_order,
            max_bins=max_bins,
            bin_period=bin_period,
            since=since,
            fill_period_bins_since=fill_period_bins_since,
            fill_end=fill_end,
        )
    if block is None:
        return None
    mat, xlabels, apps = block

    if cmap is None:
        cmap = _cadence_heatmap_colormap()

    # Larger fonts for heatmap readability in Excel.
    base_fs = 13
    fs_plot_title = int(base_fs * 1.72)
    fs_sub = int(base_fs * 1.12)
    fs_axis = int(base_fs * 1.15)
    fs_tick_y = int(base_fs * 1.08)
    fs_tick_x = int(base_fs * 1.02)
    fs_cbar = int(base_fs * 1.18)

    # Fixed requested geometry so heatmaps match across exports.
    fig, ax = plt.subplots(figsize=(16, 7))
    fig.patch.set_facecolor("#ffffff")
    ax.set_facecolor("#ffffff")

    vm = float(vmax) if vmax is not None else min(
        CADENCE_HEATMAP_VMAX_CAP,
        max(1.0, float(np.ceil(float(mat.max())))),
    )
    im = ax.imshow(mat, aspect="auto", cmap=cmap, vmin=0, vmax=vm)
    cbar_extend = "max" if float(mat.max()) > vm + 1e-9 else "neither"
    # Keep titles aligned so tight-cropping doesn't produce different PNG geometry per platform.
    ax.set_title(title, fontsize=fs_plot_title, pad=14)
    ax.text(
        0.0,
        1.01,
        subtitle,
        transform=ax.transAxes,
        ha="left",
        va="bottom",
        fontsize=fs_sub,
        color="#444444",
    )

    ax.set_xlabel("Quarter" if bin_period == "Q" else "Year" if bin_period == "Y" else "Month", fontsize=fs_axis)
    ax.set_ylabel("App", fontsize=fs_axis)
    ax.set_yticks(np.arange(len(apps)))
    ax.set_yticklabels(apps, fontsize=fs_tick_y)
    nxb = len(xlabels)
    xi = _heatmap_xtick_positions(nxb) if nxb else []
    tick_labs = [xlabels[i] for i in xi]
    ax.set_xticks(xi)
    ax.set_xticklabels(tick_labs, rotation=45, ha="right", fontsize=fs_tick_x)
    cbar = fig.colorbar(im, ax=ax, fraction=0.03, pad=0.02, extend=cbar_extend)
    cbar.set_label("Observation count (per bin)", fontsize=fs_cbar)
    nticks = min(5, max(2, int(math.ceil(vm)) + 1))
    cbar.set_ticks(np.linspace(0.0, float(vm), num=nticks))
    cbar.ax.tick_params(labelsize=int(fs_cbar * 0.9))
    # Fixed margins + no bbox_inches='tight' => consistent PNG geometry for iOS/Android.
    fig.subplots_adjust(left=0.24, right=0.985, top=0.88, bottom=0.18)
    buf = io.BytesIO()
    fig.savefig(
        buf,
        format="png",
        dpi=150,
        bbox_inches=None,
        facecolor="#ffffff",
        edgecolor="none",
        pad_inches=0.0,
    )
    plt.close(fig)
    buf.seek(0)
    return buf


def _chart_update_frequency_heatmap(version_df: pd.DataFrame) -> io.BytesIO | None:
    """
    (1) Update frequency over time by app: monthly heatmap (dated rows).
    Rows: app_name (pooled across platforms); Columns: YYYY-MM; Values: count of observations.
    """
    import matplotlib.pyplot as plt
    import numpy as np

    sub = dated_subset(version_df)
    if sub is None:
        return None
    sub = sub.copy()
    sub["_ym"] = sub["_dt"].dt.to_period("M").astype(str)
    # Pool across platforms; goal is cadence by app overall.
    piv = sub.pivot_table(index="app_name", columns="_ym", values="app_id", aggfunc="count", fill_value=0)
    if piv.empty:
        return None

    # Keep the chart readable: show at most last 30 months if very wide.
    cols = list(piv.columns)
    if len(cols) > 30:
        piv = piv[cols[-30:]]

    apps = list(piv.index.astype(str))
    xlabels = list(piv.columns.astype(str))
    mat = piv.values.astype(float)

    # Scale figure height to number of apps (cap).
    h = min(0.32 * len(apps) + 2.4, 10.5)
    w = min(0.35 * len(xlabels) + 4.8, 12.5)
    fig, ax = plt.subplots(figsize=(w, h))
    fig.patch.set_facecolor("#ffffff")
    ax.set_facecolor("#ffffff")
    vm = min(CADENCE_HEATMAP_VMAX_CAP, max(1.0, float(np.ceil(float(mat.max())))))
    im = ax.imshow(mat, aspect="auto", cmap=_cadence_heatmap_colormap(), vmin=0, vmax=vm)
    ax.set_title("1. Update frequency over time by app (monthly heatmap; dated rows)")
    ax.set_xlabel("Month")
    ax.set_ylabel("App")
    ax.set_yticks(np.arange(len(apps)))
    ax.set_yticklabels(apps, fontsize=8)
    ax.set_xticks(np.arange(len(xlabels)))
    ax.set_xticklabels(xlabels, rotation=45, ha="right", fontsize=7)
    cbar = fig.colorbar(
        im,
        ax=ax,
        fraction=0.03,
        pad=0.02,
        extend="max" if float(mat.max()) > vm + 1e-9 else "neither",
    )
    cbar.set_label("Observation count (per bin; cap 20)", fontsize=8)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(
        buf,
        format="png",
        dpi=150,
        bbox_inches="tight",
        facecolor="#ffffff",
        edgecolor="none",
        pad_inches=0.12,
    )
    plt.close(fig)
    buf.seek(0)
    return buf

def _chart_category_evolution_quartile_buckets(version_df: pd.DataFrame) -> io.BytesIO | None:
    """
    Category share shift over time (monthly, 2025–2026) for top ``update_category`` labels.

    Replaces the old 2-point quartile slope chart so volatility / intermediate fluctuations
    are visible rather than implied by straight endpoints.
    """
    import matplotlib.pyplot as plt
    import matplotlib as mpl
    import numpy as np

    import run_pipeline as rp

    sub = dated_subset(version_df)
    if sub is None or len(sub) < 8:
        return None
    # Focus window: Jul 2025 — May 2026 (no future months).
    start = pd.Timestamp("2025-07-01")
    end = pd.Timestamp("2026-05-31")
    s = sub[(sub["_dt"] >= start) & (sub["_dt"] <= end)].copy()
    if len(s) < 6:
        return None

    cat_allowed = set(rp.UPDATE_CATEGORIES)

    def _norm_cat(x: object) -> str:
        c = str(x).strip()
        return c if c in cat_allowed else "Other"

    s["_bk"] = s["update_category"].map(_norm_cat)
    s["_ym"] = s["_dt"].dt.to_period("M").astype(str)

    # Pick top-K categories within this window by overall share.
    K = 6
    overall = s["_bk"].value_counts(normalize=True)
    keys = [k for k in overall.index.tolist() if k in cat_allowed][:K]
    if not keys:
        return None

    # Monthly counts (fill missing months with 0 so volatility is visible).
    months = pd.period_range(start.to_period("M"), end.to_period("M"), freq="M").astype(str).tolist()
    counts = (
        s.pivot_table(index="_ym", columns="_bk", values="app_id", aggfunc="count", fill_value=0)
        .reindex(months, fill_value=0)
    )

    # Sparse-month gating: only plot months with ≥ 5 total observations.
    SPARSE_MIN = 5
    totals_per_month = counts.sum(axis=1)
    valid_mask = (totals_per_month >= SPARSE_MIN).values

    denom = totals_per_month.replace(0, np.nan)
    shares = (100.0 * counts.div(denom, axis=0)).fillna(0.0)
    shares = shares.reindex(columns=keys, fill_value=0.0)

    def _legend_label(full: str) -> str:
        t = full.replace(" / ", "/")
        return t if len(t) <= 44 else t[:41] + "…"

    short = [_legend_label(k) for k in keys]
    BUGFIX_KEY = "Bug fixes / performance improvements"
    BUGFIX_COLOR = "#0d3b66"  # darker blue
    tab = mpl.colormaps["tab10"]
    palette: list = []
    for i, k in enumerate(keys):
        palette.append(BUGFIX_COLOR if k == BUGFIX_KEY else tab(i % 10))

    # Smooth shares with a 3-month moving average to reduce spike noise.
    shares_sm = shares.copy()
    for k in keys:
        shares_sm[k] = shares_sm[k].astype(float).rolling(window=3, min_periods=1).mean()

    xs = np.arange(len(months))
    masked_xs = np.where(valid_mask, xs, np.nan)

    # Faceted small-multiples: one mini chart per update_category (more readable than a crowded legend).
    nkeys = len(keys)
    # Keep a 3×2 dashboard-friendly facet grid, but make it taller so panels are readable.
    ncols = 3 if nkeys >= 4 else nkeys
    nrows = int(math.ceil(nkeys / ncols))
    fig_w = 12.8 if ncols == 3 else 8.6
    fig_h = 3.6 + 3.3 * nrows
    fig, axes = plt.subplots(nrows=nrows, ncols=ncols, figsize=(fig_w, fig_h), sharex=True, sharey=True)
    fig.patch.set_facecolor("#ffffff")
    axes_arr = np.array(axes).reshape(-1)
    fs_axis = 10

    # Y zoom based on plotted (valid) months only.
    plotted_vals = shares_sm[keys].values[valid_mask]
    if plotted_vals.size:
        vmin = float(np.nanmin(plotted_vals))
        vmax = float(np.nanmax(plotted_vals))
    else:
        vmin, vmax = 0.0, 1.0
    span = max(1e-6, vmax - vmin)
    pad = max(2.0, 0.12 * span)
    y_lo = max(-0.5, vmin - pad)
    y_hi = min(100.0, vmax + pad)

    # X ticks: quarterly labels for readability.
    tick_idx = [i for i, m in enumerate(months) if m.endswith(("-01", "-04", "-07", "-10"))]
    if not tick_idx:
        tick_idx = list(range(0, len(months), 3))
    tick_lbl = [_compact_month_tick_label(months[i]) for i in tick_idx]

    # Policy / event markers (reference only; helps interpret visible shifts).
    month_to_idx = {m: i for i, m in enumerate(months)}
    refs = [
        ("2025-09", "Policy / iOS release"),
        ("2025-11", "Holiday season"),
        ("2026-01", "TikTok deadline"),
    ]
    shade_windows = [
        ("2026-01", "2026-04", "Jan–Apr 2026 window"),
    ]

    for ax_i, (k, color) in enumerate(zip(keys, palette)):
        ax = axes_arr[ax_i]
        ax.set_facecolor("#ffffff")
        ys_full = shares_sm[k].astype(float).values
        ys_masked = np.where(valid_mask, ys_full, np.nan)
        ax.plot(
            masked_xs,
            ys_masked,
            marker="o" if k == BUGFIX_KEY else None,
            markersize=3.4 if k == BUGFIX_KEY else 0,
            linewidth=2.0 if k == BUGFIX_KEY else 1.4,
            color=color,
            solid_capstyle="round",
            zorder=3,
        )
        ax.set_title(_legend_label(k), fontsize=10, fontweight="600", color="#111827", pad=6)
        ax.set_ylim(y_lo, y_hi)
        ax.set_xlim(-0.5, len(months) - 0.5)
        ax.grid(axis="y", color="#eef2f7", linestyle="-", linewidth=0.9, zorder=0)
        ax.grid(axis="x", color="#f6f8fb", linestyle="-", linewidth=0.7, zorder=0)
        for ym, _lbl in refs:
            if ym in month_to_idx:
                ax.axvline(
                    month_to_idx[ym],
                    color="#c0392b",
                    linestyle="--",
                    linewidth=1.0,
                    alpha=0.35,
                    zorder=1,
                )
        for a_ym, b_ym, _lbl in shade_windows:
            if a_ym in month_to_idx and b_ym in month_to_idx:
                a = month_to_idx[a_ym] - 0.5
                b = month_to_idx[b_ym] + 0.5
                ax.axvspan(a, b, color="#fff7ed", alpha=0.35, zorder=0)
        for side in ("top", "right"):
            ax.spines[side].set_visible(False)
        ax.spines["left"].set_color("#c7d0df")
        ax.spines["bottom"].set_color("#c7d0df")
        ax.tick_params(axis="both", colors="#4b5563", labelsize=fs_axis - 1, length=3, width=0.8)

    # Hide unused axes.
    for j in range(nkeys, len(axes_arr)):
        axes_arr[j].axis("off")

    # Only label x ticks on bottom row.
    for ax in axes_arr[max(0, nkeys - ncols) : nkeys]:
        ax.set_xticks(tick_idx)
        ax.set_xticklabels(tick_lbl, rotation=0, ha="center", fontsize=fs_axis - 1, color="#374151")

    fig.text(
        0.10,
        0.97,
        "Category share over time (monthly; 3‑month moving average) — Jan 2025 to May 2026",
        fontsize=14,
        fontweight="600",
        color="#1f2937",
        va="top",
        ha="left",
    )
    fig.text(
        0.10,
        0.915,
        "Small multiples: each panel is one update_category; excludes months with <5 observations. "
        "Dashed lines / shading are reference-only event windows.",
        fontsize=11,
        color="#5c6575",
        va="top",
        ha="left",
    )
    fig.text(
        0.10,
        0.885,
        "Date markers: 2025-09 ≈ policy/iOS-release window; 2025-11 ≈ holiday season; 2026-01 ≈ TikTok deadline. "
        "Shading highlights Jan–Apr 2026.",
        fontsize=10,
        color="#5c6575",
        va="top",
        ha="left",
    )
    fig.text(
        0.10,
        0.03,
        "Interpret platform differences cautiously (Android dated coverage is thinner).",
        fontsize=9,
        color="#6b7280",
        va="bottom",
        ha="left",
    )

    fig.subplots_adjust(left=0.07, right=0.985, top=0.86, bottom=0.10, wspace=0.22, hspace=0.45)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="#ffffff", edgecolor="none", pad_inches=0.18)
    plt.close(fig)
    buf.seek(0)
    return buf


def _classify_history_source_url(url: object) -> str:
    """Bucket ``history_source_url`` host for visualization (matches rubric single URL column)."""
    s = str(url or "").strip().lower()
    if not s:
        return "Other"
    if "web.archive.org" in s:
        return "Wayback"
    if "apkmirror.com" in s:
        return "APKMirror"
    if any(h in s for h in ("play.google.com", "apps.apple.com", "itunes.apple.com")):
        return "Store listing"
    return "Other"


def _chart_history_url_class_by_platform(version_df: pd.DataFrame) -> io.BytesIO | None:
    """Stacked 100% horizontal bars: URL host class within each platform (all observation rows)."""
    import matplotlib.pyplot as plt
    import numpy as np

    if "history_source_url" not in version_df.columns or len(version_df) < 1:
        return None

    df = version_df.copy()
    df["_ucls"] = df["history_source_url"].map(_classify_history_source_url)
    classes = ("Wayback", "APKMirror", "Store listing", "Other")
    colors = {
        "Wayback": "#5c6bc0",
        "APKMirror": "#ef6c00",
        "Store listing": "#43a047",
        "Other": "#9e9e9e",
    }

    platforms: list[str] = []
    mat: list[list[float]] = []
    for plat in ("iOS", "Android"):
        p = df[df["platform"].astype(str) == plat]
        if len(p) == 0:
            continue
        platforms.append(plat)
        n = len(p)
        mat.append([100.0 * float((p["_ucls"] == c).sum()) / float(n) for c in classes])

    if not platforms:
        return None

    mat_arr = np.array(mat, dtype=float)
    fig_h = max(2.95, 0.52 * len(platforms) + 2.05)
    fig, ax = plt.subplots(figsize=(8.0, fig_h))
    fig.patch.set_facecolor("#ffffff")
    ax.set_facecolor("#ffffff")

    y = np.arange(len(platforms))
    left = np.zeros(len(platforms))
    bar_h = 0.52
    for j, cls in enumerate(classes):
        vals = mat_arr[:, j]
        ax.barh(
            y,
            vals,
            height=bar_h,
            left=left,
            label=cls,
            color=colors[cls],
            edgecolor="#ffffff",
            linewidth=0.85,
        )
        left = left + vals

    ax.set_yticks(y)
    ax.set_yticklabels(platforms, fontsize=11, fontweight="600")
    ax.set_xlabel("Share of observations within platform (%)", fontsize=11)
    ax.set_xlim(0, 100)
    ax.set_title(
        "1C. Observation URL profile by platform (history_source_url)",
        fontsize=13,
        fontweight="600",
    )
    ax.legend(
        loc="upper center",
        bbox_to_anchor=(0.5, -0.14),
        ncol=4,
        fontsize=9,
        frameon=True,
        fancybox=True,
        edgecolor="#cdd6e4",
    )
    ax.tick_params(axis="x", labelsize=10)
    fig.subplots_adjust(left=0.14, right=0.98, top=0.88, bottom=0.22)
    buf = io.BytesIO()
    fig.savefig(
        buf,
        format="png",
        dpi=150,
        bbox_inches=None,
        facecolor="#ffffff",
        edgecolor="none",
        pad_inches=0.08,
    )
    plt.close(fig)
    buf.seek(0)
    return buf


def _chart_observation_depth_by_app_platform(version_df: pd.DataFrame) -> io.BytesIO | None:
    """
    (3) iOS vs Android observation depth: dated span in days per app per platform.
    """
    import matplotlib.pyplot as plt
    import numpy as np

    sub = dated_subset(version_df)
    if sub is None or len(sub) < 2:
        return None
    g = (
        sub.groupby(["app_name", "platform"])["_dt"]
        .agg(["min", "max", "count"])
        .reset_index()
        .rename(columns={"min": "dmin", "max": "dmax", "count": "n"})
    )
    if g.empty:
        return None
    g["span_days"] = (g["dmax"] - g["dmin"]).dt.days.astype(int)
    piv = g.pivot(index="app_name", columns="platform", values="span_days").fillna(0)
    for col in ("iOS", "Android"):
        if col not in piv.columns:
            piv[col] = 0
    piv = piv[["iOS", "Android"]]

    apps = list(piv.index.astype(str))
    ios_vals = piv["iOS"].astype(float).values
    and_vals = piv["Android"].astype(float).values

    y = np.arange(len(apps))
    n_apps = len(apps)
    bar_h = 0.34
    # Cap height so ~10 apps don’t get an oversized vertical canvas vs bar thickness.
    fig, ax = plt.subplots(figsize=(14, 6))
    fig.patch.set_facecolor("#ffffff")
    ax.set_facecolor("#ffffff")
    ax.barh(y - bar_h / 2, ios_vals, height=bar_h, label="iOS", color="#5b9bd5")
    ax.barh(y + bar_h / 2, and_vals, height=bar_h, label="Android", color="#ed7d31")
    ax.set_yticks(y)
    ax.set_yticklabels(apps, fontsize=9.0, fontweight="bold")
    ax.set_ylim(-0.62, max(n_apps - 1, 0) + 0.62)
    ax.set_xlabel("Dated span (days) = max(release_date) − min(release_date)")
    ax.set_title("iOS vs Android observation depth by app (dated span)", fontsize=14)
    ax.axvline(730, color="#666666", linestyle="--", linewidth=1.2)
    ymin, ymax = ax.get_ylim()
    ax.text(730, ymin + max(0.22, 0.04 * (ymax - ymin)), "2 years", ha="center", va="bottom", fontsize=10, color="#444444")
    ax.legend(loc="lower right", fontsize=9)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(
        buf,
        format="png",
        dpi=150,
        bbox_inches="tight",
        facecolor="#ffffff",
        edgecolor="none",
        pad_inches=0.12,
    )
    plt.close(fig)
    buf.seek(0)
    return buf


def build_explanatory_questions(version_df: pd.DataFrame) -> list[str]:
    """Mix static prompts and simple data-triggered prompts."""
    static = [
        "Which apps show accelerating vs decelerating update cadence (heatmap), and does that coincide with major platform policy events?",
        "Is Android depth shallow for specific apps (span chart) — suggesting missing captures rather than true inactivity?",
        "In the quartile category chart, do headline bucket shifts reflect disclosure strategy (templated copy) or real content change? Validate using high-confidence provenance subsets.",
    ]
    dynamic: list[str] = []
    n = len(version_df)
    if n == 0:
        return static[:3]

    ios_share = len(version_df[version_df["platform"] == "iOS"]) / n
    if ios_share > 0.72:
        dynamic.append(
            "iOS rows outweigh Android—beyond disclosure asymmetry, could scraping windows explain residual imbalance?"
        )
    elif ios_share < 0.45:
        dynamic.append(
            "Android share is high versus typical embed-vs-Play asymmetry—audit source_type per observation."
        )

    missing_ver = version_df["version_number"].map(version_string_missing).mean()
    if missing_ver > 0.45:
        dynamic.append(
            "Sparse version_number weakens semver cadence reads—rely on dated proxies and provenance filters."
        )

    sub = dated_subset(version_df)
    if sub is not None and len(sub) >= 5:
        yrs = sub["_dt"].dt.year.astype(int)
        if int(yrs.max()) == int(yrs.min()):
            dynamic.append(
                "Single-calendar-year dated span limits multi-year trend claims—expand historical captures if needed."
            )

    low_share = (version_df["confidence_level"].astype(str).str.lower() == "low").mean()
    if low_share > 0.35:
        dynamic.append(
            "Many low-confidence rows—define an analysis subset excluding feature_signal / review_inferred where "
            "appropriate."
        )

    out = static + dynamic
    seen: set[str] = set()
    uniq: list[str] = []
    for q in out:
        if q not in seen:
            seen.add(q)
            uniq.append(q)
    return uniq[:14]


def append_visualization_sheet(xlsx_path: Path, version_df: pd.DataFrame) -> None:
    """Insert / replace ``charts`` worksheet. All visuals are Matplotlib-generated PNGs embedded as images."""
    try:
        import matplotlib

        matplotlib.use("Agg")  # non-interactive backend, no display needed
        import matplotlib.pyplot as plt

        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Font
    except ImportError as e:
        raise ImportError(
            "visualization_summary requires matplotlib and openpyxl. "
            "Install: pip install matplotlib openpyxl"
        ) from e

    CHART_DIR = Path(__file__).parents[1] / "data" / "cache" / "charts"
    CHART_DIR.mkdir(parents=True, exist_ok=True)

    def save_chart_png(fig, name: str, *, dpi: int = 96, pad_inches: float = 0.03) -> Path:
        path = CHART_DIR / f"{name}.png"
        fig.savefig(
            path,
            dpi=dpi,
            bbox_inches="tight",
            pad_inches=pad_inches,
            facecolor="white",
            edgecolor="none",
        )
        plt.close(fig)
        return path

    def embed_image(ws, path: Path, anchor: str, width_pt: float, height_pt: float) -> None:
        img = XLImage(str(path))
        # openpyxl Image width/height are pixels.
        img.width = int(width_pt * 96 / 72)
        img.height = int(height_pt * 96 / 72)
        img.anchor = anchor
        ws.add_image(img)

    # --- Prepare dated data ---
    df = version_df.copy()
    df["_dt"] = parse_release_dates(df)
    dated_df_all = df[df["_dt"].notna()].copy()
    dated_df_all["_dt"] = pd.to_datetime(dated_df_all["_dt"])
    dated_df_all["month"] = dated_df_all["_dt"].dt.to_period("M").dt.to_timestamp()

    # Many Android apps have dated rows primarily before 2024; depth must use full history.
    # The 2024+ window is only for cadence/share charts to keep axes readable.
    dated_df_recent = dated_df_all[dated_df_all["month"] >= pd.Timestamp("2024-01-01")].copy()

    wb = load_workbook(xlsx_path)
    if _SHEET in wb.sheetnames:
        wb.remove(wb[_SHEET])
    ws = wb.create_sheet(_SHEET)

    # Fixed grid + anchored PNGs (more stable than one giant image)
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill

    def set_chart_grid(ws) -> None:
        for col in range(1, 20):  # A..S
            # Slightly wider grid prevents image overlap and improves readability.
            ws.column_dimensions[get_column_letter(col)].width = 6.6

        row_heights: dict[int, float] = {
            1: 20,
            2: 6,
            3: 6,
            4: 16,   # synopsis header
            5: 26,   # synopsis lines (taller so wrapped bullets aren't clipped)
            6: 26,
            7: 26,
            8: 26,
            9: 26,
            10: 26,
            11: 10,
            12: 16,
        }
        for r in range(13, 33):
            row_heights[r] = 15
        row_heights[33] = 10
        row_heights[34] = 26  # heatmap comparability note (wrapped)
        row_heights[35] = 8
        row_heights[36] = 16
        for r in range(37, 51):
            row_heights[r] = 17  # slightly taller depth chart slot
        row_heights[51] = 34  # wrapped note under chart 2 (2 lines)
        row_heights[52] = 8   # spacer row between chart 2 note and chart 3
        row_heights[53] = 16  # chart 3 title
        row_heights[54] = 22  # wrapped small-multiples descriptor
        row_heights[55] = 44  # shared legend row (chart 3)
        row_heights[56] = 6
        for r in range(57, 67):
            row_heights[r] = 15
        row_heights[67] = 8
        for r in range(68, 81):
            row_heights[r] = 15
        row_heights[81] = 10
        row_heights[82] = 10
        row_heights[83] = 6
        row_heights[84] = 6
        row_heights[85] = 6
        row_heights[86] = 14  # final note

        for row, height in row_heights.items():
            ws.row_dimensions[row].height = height

    CHART_ANCHORS: dict[str, tuple[str, float, float]] = {
        "ios_heatmap": ("A13", 320, 300),  # anchor, width_pt, height_pt
        # Shift right heatmap start to avoid overlap with the left image.
        "android_heatmap": ("K13", 320, 300),
        "depth_chart": ("A37", 570, 240),
        # Make panels slightly smaller; put shared legend on the right side.
        "event_legend": ("A55", 200, 40),
        "cat_bug": ("A57", 190, 180),
        "cat_other": ("G57", 190, 180),
        "cat_pay": ("M57", 190, 180),
        "cat_ui": ("A71", 190, 180),
        "cat_ai": ("G71", 190, 180),
        "cat_creator": ("M71", 190, 180),
    }

    def embed_chart(ws, png_path: Path, anchor: str, width_pt: float, height_pt: float) -> None:
        embed_image(ws, png_path, anchor, width_pt=width_pt, height_pt=height_pt)

    set_chart_grid(ws)

    # Text cells (written after grid, before embeds)
    fill_hdr = PatternFill(fill_type="solid", start_color="EEF2FF", end_color="EEF2FF")
    fill_note_red = PatternFill(fill_type="solid", start_color="FEE2E2", end_color="FEE2E2")
    title_font = Font(bold=True, size=12, color="111827")
    label_font = Font(bold=True, size=10, color="111827")
    note_font = Font(size=9, italic=True, color="6B7280")
    note_font_red = Font(size=9, italic=True, color="C00000")

    ws["A1"] = "Dashboard — all charts"
    ws["A1"].font = Font(bold=True, size=14, color="111827")
    ws["A1"].fill = fill_hdr
    ws.merge_cells("A1:S1")

    bullets = build_automated_trend_synopsis(version_df)
    ws["A4"] = "Automated trend synopsis"
    ws["A4"].font = label_font
    ws["A4"].fill = fill_hdr
    ws.merge_cells("A4:S4")
    syn_lines = [("• " + str(b).strip().lstrip("•").strip()) for b in bullets if str(b).strip()]
    for i in range(6):
        ws[f"A{5+i}"] = syn_lines[i] if i < len(syn_lines) else ""
        ws[f"A{5+i}"].font = Font(size=9, color="374151")
        ws[f"A{5+i}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"A{5+i}:S{5+i}")

    ws["A12"] = "1. Cadence heatmaps (monthly update frequency)"
    ws["A12"].font = title_font
    ws["A12"].fill = fill_hdr
    ws.merge_cells("A12:S12")

    ws["A34"] = (
        "Android has archive records back to 2013; heatmap clipped to iOS-comparable window (Apr 2025–May 2026). "
        "Full Android depth shown in chart 2."
    )
    ws["A34"].font = note_font
    ws["A34"].fill = fill_hdr
    ws["A34"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A34:S34")

    ws["A36"] = (
        "2. iOS vs Android dated evidence depth by app "
        "(span of collected version rows with parseable release_date)"
    )
    ws["A36"].font = title_font
    ws["A36"].fill = fill_hdr
    ws.merge_cells("A36:S36")

    ws["A51"] = (
        "Span = max(release_date) − min(release_date) across collected version-history rows only; "
        "0 means < 2 dated rows (collection depth, not app age).\n"
        "Uber Android: 302 version records; 1 dated row recovered (APKMirror returns Cloudflare 403; "
        "uploads listing page identified as future recovery path)."
    )
    ws["A51"].font = note_font_red
    ws["A51"].fill = fill_note_red
    ws["A51"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A51:S51")

    # Spacer row A52 intentionally left blank.

    ws["A53"] = "3. Category share over time (monthly; 3-month moving average)"
    ws["A53"].font = title_font
    ws["A53"].fill = fill_hdr
    ws.merge_cells("A53:S53")
    ws["A54"] = (
        "Small multiples: each panel is one update_category. Event markers are reference-only."
    )
    ws["A54"].font = Font(size=9, color="374151")
    ws["A54"].fill = fill_hdr
    ws["A54"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A54:S54")

    ws["A86"] = (
        "Dashed lines are reference-only event windows. Use as discussion guide only — not causal proof."
    )
    ws["A86"].font = note_font
    ws["A86"].fill = fill_hdr
    ws.merge_cells("A86:S86")

    # Colormap for heatmaps: 0 -> white, 1..20 -> YlOrRd
    def _heatmap_cmap_norm():
        import numpy as np
        from matplotlib.colors import BoundaryNorm, ListedColormap

        base = plt.get_cmap("YlOrRd")
        colors = ["#ffffff"] + [base(i / 20.0) for i in range(1, 21)]
        cmap = ListedColormap(colors, name="YlOrRd_0white")
        bounds = np.arange(-0.5, 21.5, 1.0)
        norm = BoundaryNorm(bounds, cmap.N, clip=True)
        return cmap, norm

    cmap, norm = _heatmap_cmap_norm()

    def _make_heatmap(platform: str, *, title: str, out_name: str) -> Path | None:
        import textwrap

        # Clip to matched window for cadence comparison (Apr 2025–May 2026).
        clip_start = pd.Timestamp("2025-04-01")
        clip_end = pd.Timestamp("2026-05-31")
        sub = dated_df_all[dated_df_all["platform"].astype(str).eq(platform)].copy()
        sub = sub[(sub["month"] >= clip_start) & (sub["month"] <= clip_end)]
        if sub.empty:
            return None
        # Ensure both platforms share the same month bins (including empty months).
        months = pd.date_range(clip_start, clip_end, freq="MS")
        piv = (
            sub.groupby(["app_name", "month"])
            .size()
            .unstack(fill_value=0)
            .sort_index(axis=0)
            .sort_index(axis=1)
        )
        piv = piv.reindex(columns=months, fill_value=0)
        anchor_key = "ios_heatmap" if str(platform) == "iOS" else "android_heatmap"
        width_pt, height_pt = CHART_ANCHORS[anchor_key][1], CHART_ANCHORS[anchor_key][2]
        fig, ax = plt.subplots(figsize=(width_pt / 72, height_pt / 72), dpi=96)
        im = ax.imshow(piv.to_numpy(), aspect="auto", cmap=cmap, norm=norm)
        ax.set_title(title, fontsize=10)
        ax.set_xlabel("")
        ax.set_ylabel("")
        ax.tick_params(axis="both", labelsize=8)
        ax.set_yticks(range(len(piv.index)))
        ax.set_yticklabels([str(x) for x in piv.index], fontsize=8)
        ax.set_xticks(range(len(piv.columns)))
        ax.set_xticklabels([pd.Timestamp(x).strftime("%b '%y") for x in piv.columns], rotation=45, ha="right", fontsize=8)
        fig.colorbar(im, ax=ax, fraction=0.05, pad=0.02)
        return save_chart_png(fig, out_name, dpi=96, pad_inches=0.02)

    def _make_depth(out_name: str) -> Path | None:
        if dated_df_all.empty:
            return None
        spans = dated_df_all.groupby(["app_name", "platform"])["_dt"].agg(["min", "max"]).reset_index()
        spans["span_days"] = (spans["max"] - spans["min"]).dt.days.astype(int)
        ios_s = spans[spans["platform"] == "iOS"][["app_name", "span_days"]].rename(columns={"span_days": "ios"})
        and_s = spans[spans["platform"] == "Android"][["app_name", "span_days"]].rename(columns={"span_days": "android"})
        merged = pd.merge(ios_s, and_s, on="app_name", how="outer").fillna(0)
        merged["android"] = merged["android"].astype(int)
        merged["ios"] = merged["ios"].astype(int)
        merged = merged.sort_values("android", ascending=False)
        width_pt, height_pt = CHART_ANCHORS["depth_chart"][1], CHART_ANCHORS["depth_chart"][2]
        fig, ax = plt.subplots(figsize=(width_pt / 72, height_pt / 72), dpi=96)
        y = range(len(merged))
        ax.barh([i + 0.18 for i in y], merged["ios"], height=0.35, color="#4472C4", label="iOS")
        ax.barh([i - 0.18 for i in y], merged["android"], height=0.35, color="#ED7D31", label="Android")
        ax.set_yticks(list(y))
        ax.set_yticklabels(merged["app_name"].tolist(), fontsize=8)
        ax.set_xlabel("Dated span in days", fontsize=8)
        ax.set_title(
            "iOS vs Android dated evidence depth by app\n"
            "(span of collected version rows with parseable release_date)",
            fontsize=10,
        )
        ax.axvline(365, color="#A3A3A3", linestyle="--", linewidth=0.8)
        ax.text(365, len(merged) - 0.5, "1 year", color="#6B7280", fontsize=8, rotation=90, va="top", ha="right")
        ax.legend(
            loc="upper right",
            fontsize=8,
            frameon=True,
            facecolor="white",
            edgecolor="#D1D5DB",
            framealpha=0.92,
        )
        ax.grid(True, axis="x", alpha=0.15)
        fig.tight_layout()
        return save_chart_png(fig, out_name, dpi=96, pad_inches=0.02)

    def _make_category_panel(category: str, *, color: str, out_name: str, show_legend: bool) -> Path | None:
        if dated_df_recent.empty:
            return None
        tot = dated_df_recent.groupby("month").size().rename("total").reset_index()
        sub = (
            dated_df_recent[dated_df_recent["update_category"].astype(str).eq(category)]
            .groupby("month")
            .size()
            .rename("n")
            .reset_index()
        )
        m = pd.merge(tot, sub, on="month", how="left").fillna({"n": 0})
        m = m[m["total"] >= 5].copy()
        if m.empty:
            return None
        m["share"] = (m["n"] / m["total"]) * 100.0
        m["share_ma3"] = m["share"].rolling(3, min_periods=1).mean()
        width_pt, height_pt = CHART_ANCHORS["cat_bug"][1], CHART_ANCHORS["cat_bug"][2]
        fig, ax = plt.subplots(figsize=(width_pt / 72, height_pt / 72), dpi=96)
        ax.plot(m["month"], m["share_ma3"], color=color, linewidth=1.3)
        # Reference markers (colored).
        ref = [
            (pd.Timestamp("2025-09-01"), "Sep 2025 — iOS 18 / SDK deadline", "#2563EB"),
            (pd.Timestamp("2025-11-01"), "Nov 2025 — holiday commercial window", "#16A34A"),
            (pd.Timestamp("2026-01-01"), "Jan 2026 — TikTok US deadline", "#7C3AED"),
        ]
        for dtx, _lab, col in ref:
            ax.axvline(dtx, color=col, linestyle="--", linewidth=1.1, alpha=0.85)
        ax.set_ylim(0, 70)
        ax.set_title(category, fontsize=8.5, fontweight="normal")
        ax.set_ylabel("Share (%)", fontsize=7.5)
        ax.tick_params(axis="both", labelsize=7)
        # Quarterly ticks (every 3 months) for readability.
        start = pd.Timestamp(m["month"].min()).to_period("Q").start_time
        end = pd.Timestamp(m["month"].max()).to_period("Q").start_time
        ticks = list(pd.date_range(start, end, freq="QS"))
        ax.set_xticks(ticks)
        ax.set_xticklabels([t.strftime("%b '%y") for t in ticks], rotation=45, ha="right")
        # One shared legend: render it once (first panel), not repeated.
        if show_legend:
            from matplotlib.lines import Line2D

            line_handles = [Line2D([0], [0], color=c, linestyle="--", linewidth=1.1) for _, _, c in ref]
            ax.legend(
                handles=line_handles,
                labels=[lab for _, lab, _ in ref],
                loc="upper left",
                fontsize=6.5,
                frameon=True,
                facecolor="white",
                edgecolor="#E5E7EB",
                framealpha=0.92,
                borderpad=0.35,
                labelspacing=0.25,
                handlelength=2.2,
                handletextpad=0.5,
            )
        ax.grid(True, axis="y", alpha=0.15)
        fig.tight_layout()
        return save_chart_png(fig, out_name, dpi=96, pad_inches=0.02)

    def _make_event_legend(out_name: str) -> Path:
        """Standalone legend PNG so small-multiple panels stay uncluttered."""
        import matplotlib.pyplot as plt
        from matplotlib.lines import Line2D

        width_pt, height_pt = CHART_ANCHORS["event_legend"][1], CHART_ANCHORS["event_legend"][2]
        fig, ax = plt.subplots(figsize=(width_pt / 72, height_pt / 72), dpi=96)
        ax.axis("off")
        ref = [
            ("Sep 2025 — iOS 18 / App Store SDK deadline window", "#2563EB"),
            ("Nov 2025 — holiday commercial window", "#16A34A"),
            ("Jan 2026 — TikTok US deadline", "#7C3AED"),
        ]
        handles = [Line2D([0], [0], color=c, linestyle="--", linewidth=1.4) for _, c in ref]
        labels = [lab for lab, _c in ref]
        ax.legend(
            handles=handles,
            labels=labels,
            loc="upper left",
            fontsize=8,
            frameon=True,
            facecolor="white",
            edgecolor="#E5E7EB",
            framealpha=0.95,
            borderpad=0.6,
            labelspacing=0.8,
            handlelength=3.0,
            handletextpad=0.8,
        )
        fig.tight_layout(pad=0.2)
        return save_chart_png(fig, out_name, dpi=96, pad_inches=0.02)

    # Generate + embed in order (exact requested)
    ios_png = _make_heatmap("iOS", title="Cadence heatmap (monthly) — iOS", out_name="heatmap_ios_grid")
    if ios_png:
        a, w, h = CHART_ANCHORS["ios_heatmap"]
        embed_chart(ws, ios_png, a, w, h)
    and_png = _make_heatmap("Android", title="Cadence heatmap (monthly) — Android", out_name="heatmap_android_grid")
    if and_png:
        a, w, h = CHART_ANCHORS["android_heatmap"]
        embed_chart(ws, and_png, a, w, h)

    depth_png = _make_depth("depth_grid")
    if depth_png:
        a, w, h = CHART_ANCHORS["depth_chart"]
        embed_chart(ws, depth_png, a, w, h)

    cat_colors = {
        "Bug fixes / performance improvements": "#1F3864",
        "Other": "#ED7D31",
        "Payments / monetization": "#375623",
        "UI / design changes": "#C00000",
        "AI-related features": "#7030A0",
        "Creator tools / content features": "#833C00",
    }
    panels = [
        ("cat_bug", "Bug fixes / performance improvements"),
        ("cat_other", "Other"),
        ("cat_pay", "Payments / monetization"),
        ("cat_ui", "UI / design changes"),
        ("cat_ai", "AI-related features"),
        ("cat_creator", "Creator tools / content features"),
    ]
    for key, cat in panels:
        png = _make_category_panel(
            cat,
            color=cat_colors[cat],
            out_name=f"panel_{key}",
            show_legend=False,
        )
        if png:
            a, w, h = CHART_ANCHORS[key]
            embed_chart(ws, png, a, w, h)

    # Shared legend PNG (one time, bottom)
    legend_png = _make_event_legend("event_markers_legend")
    a, w, h = CHART_ANCHORS["event_legend"]
    embed_chart(ws, legend_png, a, w, h)

    wb.save(xlsx_path)

    # CLEANUP: remove temporary PNGs after embedding.
    for p in CHART_DIR.glob("*.png"):
        try:
            p.unlink()
        except OSError:
            pass

    return


    def _heatmap(ax, piv: pd.DataFrame, title: str) -> None:
        mat = piv.to_numpy()
        im = ax.imshow(mat, aspect="auto", cmap="YlOrRd", vmin=0, vmax=20)
        ax.set_title(title, fontsize=10)
        ax.set_xlabel("")
        ax.set_ylabel("")
        ax.set_yticks(range(len(piv.index)))
        ax.set_yticklabels([str(x) for x in piv.index], fontsize=8)
        ax.set_xticks(range(len(piv.columns)))
        ax.set_xticklabels(
            [pd.Timestamp(x).strftime("%b '%y") for x in piv.columns],
            rotation=45,
            ha="right",
            fontsize=8,
        )
        # Compact colorbar.
        cb = ax.figure.colorbar(im, ax=ax, fraction=0.05, pad=0.02)
        cb.ax.tick_params(labelsize=8)

    # CHART 1 — iOS cadence heatmap
    ios = dated_df_recent[dated_df_recent["platform"].astype(str).eq("iOS")].copy()
    if not ios.empty:
        piv = (
            ios.groupby(["app_name", "month"])
            .size()
            .unstack(fill_value=0)
            .sort_index(axis=0)
            .sort_index(axis=1)
        )
        fig, ax = plt.subplots(figsize=(6.0, 3.7))
        _heatmap(ax, piv, "Cadence heatmap (monthly) — iOS")
        p = save_chart_png(fig, "cadence_heatmap_ios")
        embed_image(ws, p, anchor="A20", width_pt=470, height_pt=340)

    # CHART 2 — Android cadence heatmap
    andr = dated_df_recent[dated_df_recent["platform"].astype(str).eq("Android")].copy()
    if not andr.empty:
        piv = (
            andr.groupby(["app_name", "month"])
            .size()
            .unstack(fill_value=0)
            .sort_index(axis=0)
            .sort_index(axis=1)
        )
        fig, ax = plt.subplots(figsize=(6.0, 3.7))
        _heatmap(ax, piv, "Cadence heatmap (monthly) — Android")
        p = save_chart_png(fig, "cadence_heatmap_android")
        embed_image(ws, p, anchor="H20", width_pt=470, height_pt=340)

    # (Moved into PNG annotation card below for layout stability.)

    # CHART 3 — observation depth bar chart
    if not dated_df_all.empty:
        spans = (
            dated_df_all.groupby(["app_name", "platform"])["_dt"]
            .agg(["min", "max"])
            .reset_index()
        )
        spans["span_days"] = (spans["max"] - spans["min"]).dt.days.astype(int)
        ios_s = spans[spans["platform"] == "iOS"][["app_name", "span_days"]].rename(columns={"span_days": "ios"})
        and_s = spans[spans["platform"] == "Android"][["app_name", "span_days"]].rename(columns={"span_days": "android"})
        merged = pd.merge(ios_s, and_s, on="app_name", how="outer").fillna(0)
        merged["android"] = merged["android"].astype(int)
        merged["ios"] = merged["ios"].astype(int)
        merged = merged.sort_values("android", ascending=False)

        fig, ax = plt.subplots(figsize=(10.0, 3.1))
        y = range(len(merged))
        ax.barh([i + 0.18 for i in y], merged["ios"], height=0.35, color="#4472C4", label="iOS")
        ax.barh([i - 0.18 for i in y], merged["android"], height=0.35, color="#ED7D31", label="Android")
        ax.set_yticks(list(y))
        ax.set_yticklabels(merged["app_name"].tolist(), fontsize=8)
        ax.set_xlabel("Dated span in days", fontsize=8)
        ax.set_title("iOS vs Android observation depth by app (dated span)", fontsize=10)
        ax.axvline(365, color="#A3A3A3", linestyle="--", linewidth=0.8)
        ax.text(365, len(merged) - 0.5, "1 year", color="#6B7280", fontsize=8, rotation=90, va="top", ha="right")
        ax.legend(loc="lower right", fontsize=8, frameon=False)
        fig.tight_layout()
        p = save_chart_png(fig, "observation_depth_by_app")
        embed_image(ws, p, anchor="A44", width_pt=980, height_pt=300)

    # CHART 4-9 — category share small multiples (combined 3×2)
    cats = [
        "Bug fixes / performance improvements",
        "Other",
        "Payments / monetization",
        "UI / design changes",
        "AI-related features",
        "Creator tools / content features",
    ]
    colors = {
        "Bug fixes / performance improvements": "#1F3864",
        "Other": "#ED7D31",
        "Payments / monetization": "#375623",
        "UI / design changes": "#C00000",
        "AI-related features": "#7030A0",
        "Creator tools / content features": "#833C00",
    }

    if not dated_df_recent.empty and "update_category" in dated_df_recent.columns:
        tot = dated_df_recent.groupby("month").size().rename("total").reset_index()
        fig, axes = plt.subplots(2, 3, figsize=(13.0, 6.2), sharex=True, sharey=True)
        axes = axes.flatten()
        for i, cat in enumerate(cats):
            ax = axes[i]
            sub = (
                dated_df_recent[dated_df_recent["update_category"].astype(str).eq(cat)]
                .groupby("month")
                .size()
                .rename("n")
                .reset_index()
            )
            m = pd.merge(tot, sub, on="month", how="left").fillna({"n": 0})
            m = m[m["total"] >= 5].copy()
            if m.empty:
                ax.set_title(cat, fontsize=10, fontweight="bold")
                ax.set_ylim(0, 70)
                ax.grid(True, axis="y", alpha=0.15)
                continue
            m["share"] = (m["n"] / m["total"]) * 100.0
            m["share_ma3"] = m["share"].rolling(3, min_periods=1).mean()
            ax.plot(m["month"], m["share_ma3"], color=colors.get(cat, "#111827"), linewidth=2.2)
            for dt in [pd.Timestamp("2025-09-01"), pd.Timestamp("2025-11-01"), pd.Timestamp("2026-01-01")]:
                ax.axvline(dt, color="#BFBFBF", linestyle="--", linewidth=0.9)
            ax.axvspan(pd.Timestamp("2026-01-01"), pd.Timestamp("2026-04-01"), color="#ED7D31", alpha=0.08)
            ax.set_title(cat, fontsize=10, fontweight="bold")
            ax.set_ylim(0, 70)
            ax.grid(True, axis="y", alpha=0.15)
            ax.tick_params(axis="both", labelsize=9)
            ticks = m["month"].tolist()[:: max(1, int(len(m) / 6))]
            ax.set_xticks(ticks)
            ax.set_xticklabels([pd.Timestamp(ts).strftime("%b '%y") for ts in ticks], rotation=45, ha="right")

        fig.suptitle("Category share small multiples (monthly; 3-month MA)", fontsize=12, fontweight="bold")
        fig.text(
            0.5,
            0.01,
            "Excludes months with <5 observations. Dashed lines are reference-only event windows. "
            "Interpret platform differences cautiously (Android dated coverage is thinner).",
            ha="center",
            fontsize=9,
            color="#6B7280",
            style="italic",
        )
        fig.tight_layout(rect=[0, 0.04, 1, 0.95])
        p = save_chart_png(fig, "category_small_multiples_3x2")
        p_note = _annotation_card_png()
        embed_image(ws, p_note, anchor="A74", width_pt=980, height_pt=160)
        embed_image(ws, p, anchor="A80", width_pt=980, height_pt=520)

    wb.save(xlsx_path)

    # CLEANUP: remove temporary PNGs after embedding.
    for p in CHART_DIR.glob("*.png"):
        try:
            p.unlink()
        except OSError:
            pass


def try_append_visualization_sheet(xlsx_path: Path, version_df: pd.DataFrame) -> None:
    try:
        append_visualization_sheet(xlsx_path, version_df)
    except Exception as e:
        print(f"[warn] charts sheet skipped: {e}", file=sys.stderr)
